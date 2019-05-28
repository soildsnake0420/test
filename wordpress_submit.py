# 本地Chrome浏览器的静默默模式设置：
from selenium import webdriver  # 从selenium库中调用webdriver模块
from selenium.webdriver.chrome.options import Options  # 从options模块中调用Options类
from openpyxl import load_workbook
import time, openpyxl

userid = 'user_login'
userps = 'user_pass'
submits = 'wp-submit'

#输入用户名
username = '' 

#输入密码
password = '' 

#输入的地址，XXX为你的blog域名
post_url = 'http://XXX.com/wp-admin/post-new.php'

#存有内容的 xlsx 文件。 排列规则第一行为标题，比如：A1，B1，C1 ，第二行为内容，比如：A2，B2，C2. A1 A2 为一个题目的标题和内容。
excelfile = 'novel.xlsx'
excelsheet = 'sheet1'




#提交一篇文章

class wordpress_submit():


    def __init__(self, excelfile, excelsheet, post_url, username, password):

        # 打开文件
        self.wb = load_workbook(excelfile)
        # 获取工作表
        self.ws = self.wb[excelsheet]

        #遍历所有标题和内容，并将只组合起来。
        for t,c in zip(self.load_title(), self.load_conent()):
            
            #打开地址
            self.open_url(post_url)
            
            #输入用户名邮箱
            self.login_user(username, password)
            
            #将标题和内容输入进去并提交
            self.sumbit_post(t, c)
            
            #等待两秒之后再继续。因为提交需要时间
            time.sleep(2)

    #----- 以下为方法 ------#
    
    # 获取标题
    def load_title(self):

        noveltitle = []

        # 将文件中的标题读取
        for row_cell in self.ws['A1':'C1']:
            for cell in row_cell:
                title = cell.value
                noveltitle.append(title)

        # noveltitle为列表，要取出值必须用循环
        return noveltitle

    # 获取内容
    def load_conent(self):

        # 文章内容列表
        novelconent = []

        # 将文件中的内容读取
        for row_cel in self.ws['A2':'C2']:
            for cell in row_cel:
                conent = cell.value
                novelconent.append(conent)

        # novelconent为列表，要取出值必须用循环
        return novelconent






    def open_url(self,login_url):

        #------ 浏览器后台运行 --------#
        chrome_options = Options()  # 实例化Option对象
        chrome_options.add_argument('--headless')  # 把Chrome浏览器设置为静默模式
        self.driver = webdriver.Chrome(options=chrome_options)  # 设置引擎为Chrome，在后台默默运行
        
        #----- 浏览器前台运行，用与测试 -----#
        # 设置chrome浏览器
        #self.driver = webdriver.Chrome()

        # 打开网址
        self.driver.get(login_url)



    def login_user(self, username, password):

        # 输入账号
        usernames = self.driver.find_element_by_id(userid)
        usernames.send_keys(username)

        # 输入密码
        passwords = self.driver.find_element_by_id(userps)
        passwords.send_keys(password)

        # 登录
        enter = self.driver.find_element_by_id(submits)
        enter.click()

        # 等待时间，此处为最长等待时间，如果15秒内加载完成，会自动结束。
        self.driver.implicitly_wait(15)


    #提交标题和内容
    def sumbit_post(self, post_title, post_conent):

        post_titles = self.driver.find_element_by_id('post-title-0')
        post_titles.send_keys(post_title)

        post_conents = self.driver.find_element_by_id('post-content-0')
        post_conents.send_keys(post_conent)

        post_submit = self.driver.find_element_by_xpath('//*[@id="editor"]/div/div/div/div[1]/div[2]/button[2]')
        post_submit.click()


        post_sbb = self.driver.find_element_by_xpath('//*[@id="editor"]/div/div/div/div[3]/div/div/div[1]/div/button')
        
        #由于wordpress同一文章提交要2次，所以click要使用2次，
        post_sbb.click()

        self.driver.implicitly_wait(15)

        post_sbb.click()

        time.sleep(2)
        #一切完毕后关闭浏览器，否则会导致网址错误。
        self.driver.close()


wds = wordpress_submit(excelfile, excelsheet, post_url, username, password)
